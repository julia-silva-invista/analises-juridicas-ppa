# -*- coding: utf-8 -*-
"""
Coleta de Informações — preenche planilha x.xlsx com dados da Predictus
"""

import os
import re
import shutil
import unicodedata
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
from typing import Optional, Tuple, Any

from dossie_ppa import preencher_passivo_dossie, preencher_ativos_e_passivo_dossie, _fmt_valor_br

# ── Constantes ───────────────────────────────────────────────────────────────

TEMPLATE_PATH = "x.xlsx"

CLASSE_MAP = {
    "EXECUCAO DE TITULO EXTRAJUDICIAL": "Execução de Título Extrajudicial",
    "EMBARGOS A EXECUCAO": "Embargos à Execução",
    "EMBARGOS À EXECUÇÃO": "Embargos à Execução",
    "CARTA PRECATORIA CIVEL": "Carta Precatória Cível",
    "CARTA PRECATÓRIA CÍVEL": "Carta Precatória Cível",
    "AGRAVO DE INSTRUMENTO": "Agravo de Instrumento",
    "ACAO TRABALHISTA - RITO ORDINARIO": "Ação Trabalhista — Rito Ordinário",
    "AÇÃO TRABALHISTA - RITO ORDINÁRIO": "Ação Trabalhista — Rito Ordinário",
    "ACAO TRABALHISTA - RITO SUMARIÍSSIMO": "Ação Trabalhista — Rito Sumaríssimo",
    "ACAO TRABALHISTA - RITO SUMARÍSSIMO": "Ação Trabalhista — Rito Sumaríssimo",
    "ACAO CIVIL PUBLICA": "Ação Civil Pública",
    "AÇÃO CIVIL PÚBLICA": "Ação Civil Pública",
    "EXECUCAO FISCAL": "Execução Fiscal",
    "MANDADO DE SEGURANCA": "Mandado de Segurança",
    "MANDADO DE SEGURANÇA": "Mandado de Segurança",
    "ACAO DECLARATORIA": "Ação Declaratória",
    "AÇÃO DECLARATÓRIA": "Ação Declaratória",
    "MONITORIA": "Monitória",
    "MONITÓRIA": "Monitória",
    "CUMPRIMENTO DE SENTENCA": "Cumprimento de Sentença",
    "CUMPRIMENTO DE SENTENÇA": "Cumprimento de Sentença",
    "ACAO DE COBRANCA": "Ação de Cobrança",
    "AÇÃO DE COBRANÇA": "Ação de Cobrança",
    "APELACAO CIVEL": "Apelação Cível",
    "APELAÇÃO CÍVEL": "Apelação Cível",
    "EMBARGOS DE DECLARACAO": "Embargos de Declaração",
    "EMBARGOS DE DECLARAÇÃO": "Embargos de Declaração",
    "RECURSO DE REVISTA": "Recurso de Revista",
    "RECURSO ORDINARIO TRABALHISTA": "Recurso Ordinário Trabalhista",
    "RECURSO ORDINÁRIO TRABALHISTA": "Recurso Ordinário Trabalhista",
    "ACAO RESCISORIA": "Ação Rescisória",
    "AÇÃO RESCISÓRIA": "Ação Rescisória",
    "RECLAMACAO TRABALHISTA": "Reclamação Trabalhista",
    "RECLAMAÇÃO TRABALHISTA": "Reclamação Trabalhista",
    "DISSIDIO INDIVIDUAL TRABALHISTA": "Dissídio Individual Trabalhista",
    "DISSÍDIO INDIVIDUAL TRABALHISTA": "Dissídio Individual Trabalhista",
    "ACAO DE INDENIZACAO POR DANO MORAL": "Ação de Indenização por Dano Moral",
    "AÇÃO DE INDENIZAÇÃO POR DANO MORAL": "Ação de Indenização por Dano Moral",
}

STATUS_MAP = {
    "EM TRAMITACAO": "Em tramitação",
    "EM TRAMITAÇÃO": "Em tramitação",
    "ARQUIVADO": "Arquivado definitivamente",
    "ARQUIVADO DEFINITIVAMENTE": "Arquivado definitivamente",
    "ARQUIVAMENTO DEFINITIVO": "Arquivado definitivamente",
    "BAIXADO": "Arquivado definitivamente",
    "ARQUIVAMENTO": "Arquivamento",
    "ARQUIVAMENTO PROVISORIO": "Arquivamento provisório",
    "ARQUIVAMENTO PROVISÓRIO": "Arquivamento provisório",
    "SUSPENSO": "Suspenso",
    "EXTINTO": "Extinto",
    "JULGADO": "Julgado",
    "TRANSITADO EM JULGADO": "Transitado em julgado",
}

_LOWERCASE_WORDS = {
    "de", "da", "do", "das", "dos", "e", "em", "a", "o", "ao", "à",
    "na", "no", "nas", "nos", "por", "para", "com", "sem", "sob",
    "sobre", "entre",
}

# ── Helpers ───────────────────────────────────────────────────────────────────

def _title_case(text: str) -> str:
    if not text:
        return ""
    words = str(text).strip().split()
    result = []
    for i, word in enumerate(words):
        if i == 0 or word.lower() not in _LOWERCASE_WORDS:
            result.append(word.capitalize())
        else:
            result.append(word.lower())
    return " ".join(result)


def _format_document(raw: str) -> str:
    d = re.sub(r"\D", "", raw)
    if len(d) == 11:
        return f"{d[:3]}.{d[3:6]}.{d[6:9]}-{d[9:]}"
    if len(d) == 14:
        return f"{d[:2]}.{d[2:5]}.{d[5:8]}/{d[8:12]}-{d[12:]}"
    return raw


def _doc_label(raw: str) -> str:
    d = re.sub(r"\D", "", raw)
    if len(d) == 11:
        return "CPF"
    if len(d) == 14:
        return "CNPJ"
    return ""


def _format_parties(party_str: str) -> str:
    raw = str(party_str).strip()
    if raw in ("", "nan", "None"):
        return ""
    parties = []
    for part in raw.split("|"):
        part = part.strip()
        if not part:
            continue
        if ":" in part:
            part = part.split(":", 1)[1].strip()
        m = re.search(r"\[(\d+)\]", part)
        name = re.sub(r"\[.*?\]", "", part).strip()
        name = _title_case(name)
        if not name:
            continue
        if m:
            doc_raw = m.group(1)
            label = _doc_label(doc_raw)
            doc_fmt = _format_document(doc_raw)
            if label:
                parties.append(f"{name} ({label} nº {doc_fmt})")
            else:
                parties.append(name)
        else:
            parties.append(name)
    return "; ".join(parties)


def _format_classe(classe: str) -> str:
    s = str(classe).strip()
    if s in ("", "nan", "None"):
        return ""
    return CLASSE_MAP.get(s.upper(), _title_case(s))


def _format_status(status: str) -> str:
    s = str(status).strip()
    if s in ("", "nan", "None"):
        return ""
    return STATUS_MAP.get(s.upper(), _title_case(s))


def _parse_date(val) -> Optional[str]:
    s = str(val).strip()
    if s in ("", "nan", "None", "NaT"):
        return None
    if re.match(r"\d{2}/\d{2}/\d{4}", s):
        return s
    try:
        from dateutil import parser as dp
        return dp.parse(s, dayfirst=True).strftime("%d/%m/%Y")
    except Exception:
        return s


def _is_trabalhista(ramo: str) -> bool:
    r = str(ramo).upper()
    return any(kw in r for kw in ("TRABALHO", "TRABALHISTA"))


def _is_fiscal(ramo: str, classe: str) -> bool:
    s = (str(ramo) + " " + str(classe)).upper()
    return any(kw in s for kw in (
        "FISCAL", "TRIBUT", "DIVIDA ATIVA", "DÍVIDA ATIVA", "EXECUCAO FISCAL", "EXECUÇÃO FISCAL",
    ))


def _get_nome_from_filename(filepath: str) -> str:
    base = os.path.splitext(os.path.basename(filepath))[0]
    return re.sub(r"^Predictus\s*[-–—]\s*", "", base, flags=re.IGNORECASE).strip()


def _first_empty_row(ws, col: int = 1) -> int:
    for r in range(2, ws.max_row + 2):
        if ws.cell(row=r, column=col).value is None:
            return r
    return ws.max_row + 1


def _col(df: pd.DataFrame, keywords: list) -> str:
    for kw in keywords:
        for col in df.columns:
            if kw.lower() in col.lower():
                return col
    return ""


# ── Função principal ──────────────────────────────────────────────────────────

def coleta_gerar(excel_files) -> Tuple[str, str, Any]:
    """
    Lê Excel(s) da Predictus e preenche o template x.xlsx.
    Retorna: (log_text, status_msg, caminho_do_arquivo_ou_None)
    """
    lines: list = []

    def log(msg: str):
        lines.append(msg)
        return "\n".join(lines)

    if not excel_files:
        return "Nenhum arquivo enviado.", "Erro: nenhum arquivo", None

    if not os.path.exists(TEMPLATE_PATH):
        return (
            f"Template '{TEMPLATE_PATH}' não encontrado.\n"
            "Certifique-se de que o arquivo x.xlsx está na mesma pasta que os arquivos .py.",
            "Erro: template ausente",
            None,
        )

    os.makedirs("resultados", exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = f"resultados/coleta_{ts}.xlsx"
    shutil.copy2(TEMPLATE_PATH, out_path)

    wb = load_workbook(out_path)
    ws_trab   = wb["Trabalhista"]
    ws_fiscal = wb["Fiscal & Cível"]
    ws_capa   = wb["Capa"]

    row_trab   = _first_empty_row(ws_trab)
    row_fiscal = _first_empty_row(ws_fiscal)
    row_capa   = _first_empty_row(ws_capa)

    total_trab = total_fiscal = 0

    for file_obj in excel_files:
        fp = file_obj.name if hasattr(file_obj, "name") else str(file_obj)
        log(f"📂 {os.path.basename(fp)}")

        try:
            xl = pd.ExcelFile(fp)
            sheet = "Dossiê Jurídico" if "Dossiê Jurídico" in xl.sheet_names else xl.sheet_names[0]
            df = pd.read_excel(fp, sheet_name=sheet, dtype=str)
        except Exception as exc:
            log(f"   ❌ Erro ao ler arquivo: {exc}")
            continue

        if df.empty:
            log("   ⚠️ Planilha vazia.")
            continue

        df.columns = [c.strip() for c in df.columns]

        nome = _get_nome_from_filename(fp)
        cpf_col = _col(df, ["termo buscado", "termo", "buscado", "cpf", "cnpj"])
        cpf = str(df[cpf_col].iloc[0]) if cpf_col else ""

        log(f"   👤 {nome}  |  {cpf}")
        log(f"   📋 {len(df)} processo(s)")

        ws_capa.cell(row=row_capa, column=1, value=nome)
        ws_capa.cell(row=row_capa, column=3, value=cpf)
        row_capa += 1

        col_ramo    = _col(df, ["ramo do direito", "ramo"])
        col_valor   = _col(df, ["valor da causa", "valor"])
        col_cnj     = _col(df, ["n° processo", "nº processo", "numero", "processo"])
        col_data    = _col(df, ["data de distribui", "distribuição", "distribuicao"])
        col_classe  = _col(df, ["classe processual", "classe"])
        col_ativo   = _col(df, ["partes ativas", "polo ativo"])
        col_passivo = _col(df, ["partes passivas", "polo passivo"])
        col_status  = _col(df, ["status", "situação", "situacao"])

        n_t = n_f = 0

        for _, row in df.iterrows():

            def g(col_name: str) -> str:
                if not col_name:
                    return ""
                v = row.get(col_name, "")
                return str(v) if v and str(v) not in ("nan", "None") else ""

            ramo = g(col_ramo)
            valor_raw = g(col_valor)
            try:
                valor = float(valor_raw.replace(",", ".")) if valor_raw else None
            except Exception:
                valor = None

            proc = {
                "cnj":     g(col_cnj),
                "vinc":    nome,
                "data":    _parse_date(g(col_data)),
                "classe":  _format_classe(g(col_classe)),
                "valor":   valor,
                "ativo":   _format_parties(g(col_ativo)),
                "passivo": _format_parties(g(col_passivo)),
                "status":  _format_status(g(col_status)),
            }

            if _is_trabalhista(ramo):
                r, ws = row_trab, ws_trab
                row_trab += 1
                n_t += 1
            else:
                r, ws = row_fiscal, ws_fiscal
                row_fiscal += 1
                n_f += 1

            ws.cell(row=r, column=1, value=proc["cnj"])
            ws.cell(row=r, column=2, value=proc["vinc"])
            ws.cell(row=r, column=3, value=proc["data"])
            ws.cell(row=r, column=4, value=proc["classe"])
            ws.cell(row=r, column=5, value=proc["valor"])
            # coluna 6 = Saldo Atualizado (fórmula — não sobrescrever)
            ws.cell(row=r, column=7, value=proc["ativo"])
            ws.cell(row=r, column=8, value=proc["passivo"])
            ws.cell(row=r, column=9, value=proc["status"])

        total_trab  += n_t
        total_fiscal += n_f
        log(f"   ✅ {n_t} trabalhista(s) | {n_f} fiscal/cível")

    wb.save(out_path)

    total = total_trab + total_fiscal
    log(f"\n✅ Concluído! {total} processo(s) preenchido(s):")
    log(f"   • Trabalhista:    {total_trab}")
    log(f"   • Fiscal & Cível: {total_fiscal}")

    status_msg = f"{total} processos preenchidos — {total_trab} trabalhistas, {total_fiscal} fiscal/cível"
    return "\n".join(lines), status_msg, out_path


# ── Dossiê atualizado: passivo (Seção 3) a partir da Predictus ────────────────

def _classificar_predictus(excel_files, log) -> Tuple[list, list, list]:
    """Lê os Excel(s) da Predictus e separa os processos em fiscal / trabalhista / cível."""
    fiscais, trabalhistas, civeis = [], [], []
    for file_obj in excel_files:
        fp = file_obj.name if hasattr(file_obj, "name") else str(file_obj)
        log(f"📂 {os.path.basename(fp)}")
        try:
            xl = pd.ExcelFile(fp)
            sheet = "Dossiê Jurídico" if "Dossiê Jurídico" in xl.sheet_names else xl.sheet_names[0]
            df = pd.read_excel(fp, sheet_name=sheet, dtype=str)
        except Exception as exc:
            log(f"   ❌ Erro ao ler arquivo: {exc}")
            continue
        if df.empty:
            log("   ⚠️ Planilha vazia.")
            continue
        df.columns = [c.strip() for c in df.columns]
        nome = _get_nome_from_filename(fp)

        col_ramo    = _col(df, ["ramo do direito", "ramo"])
        col_valor   = _col(df, ["valor da causa", "valor"])
        col_cnj     = _col(df, ["n° processo", "nº processo", "numero", "processo"])
        col_data    = _col(df, ["data de distribui", "distribuição", "distribuicao"])
        col_classe  = _col(df, ["classe processual", "classe"])
        col_ativo   = _col(df, ["partes ativas", "polo ativo"])
        col_passivo = _col(df, ["partes passivas", "polo passivo"])
        col_status  = _col(df, ["status", "situação", "situacao"])

        n_f = n_t = n_c = 0
        for _, row in df.iterrows():
            def g(col_name: str) -> str:
                if not col_name:
                    return ""
                v = row.get(col_name, "")
                return str(v) if v and str(v) not in ("nan", "None") else ""

            ramo = g(col_ramo)
            classe = g(col_classe)
            valor_raw = g(col_valor)
            try:
                valor = float(valor_raw.replace(",", ".")) if valor_raw else None
            except Exception:
                valor = None

            proc = {
                "cnj":     g(col_cnj),
                "vinc":    nome,
                "passivo": _format_parties(g(col_passivo)),
                "data":    _parse_date(g(col_data)) or "",
                "ativo":   _format_parties(g(col_ativo)),
                "valor":   valor,
                "status":  _format_status(g(col_status)),
            }
            if _is_trabalhista(ramo):
                trabalhistas.append(proc); n_t += 1
            elif _is_fiscal(ramo, classe):
                fiscais.append(proc); n_f += 1
            else:
                civeis.append(proc); n_c += 1
        log(f"   👤 {nome} — {n_f} fiscal | {n_t} trabalhista | {n_c} cível")
    return fiscais, trabalhistas, civeis


def coleta_gerar_dossie(excel_files, dossie_file):
    """
    Preenche a Seção 3 (Passivo) de um dossiê PPA com os processos da Predictus.
    Retorna: (log_text, status_msg, caminho_do_docx_ou_None).
    """
    lines: list = []

    def log(msg: str):
        lines.append(msg)
        return "\n".join(lines)

    if not excel_files:
        return "Nenhum Excel da Predictus enviado.", "Erro: nenhum Excel da Predictus", None

    try:
        fiscais, trabalhistas, civeis = _classificar_predictus(excel_files, log)
    except Exception as exc:
        return log(f"\n❌ Erro ao ler a Predictus: {exc}"), "Erro na leitura da Predictus", None

    total = len(fiscais) + len(trabalhistas) + len(civeis)
    if total == 0:
        return log("\n⚠️ Nenhum processo identificado nos arquivos."), "Nenhum processo identificado", None

    dossie_path = None
    if dossie_file:
        dossie_path = dossie_file.name if hasattr(dossie_file, "name") else str(dossie_file)
        log(f"\n📄 Atualizando dossiê enviado: {os.path.basename(dossie_path)}")
    else:
        log("\n📄 Nenhum dossiê enviado — gerando esqueleto com o passivo preenchido.")

    try:
        out = preencher_passivo_dossie(dossie_path, fiscais, trabalhistas, civeis)
    except Exception as exc:
        return log(f"\n❌ Erro ao preencher o dossiê: {exc}"), "Erro ao preencher o dossiê", None

    log(f"\n✅ Passivo preenchido: {len(fiscais)} fiscal · {len(trabalhistas)} trabalhista · {len(civeis)} cível.")
    status = f"Passivo preenchido — {len(fiscais)} fiscal, {len(trabalhistas)} trabalhista, {len(civeis)} cível"
    return "\n".join(lines), status, out


# ── Excel "Coleta de Informações sobre Caso" (Matrículas + Fiscal & Cível + Trabalhista + E-cac) ──
# Formato próprio (não Predictus): abas dedicadas já vêm com colunas limpas, sem precisar
# adivinhar por palavra-chave. Preenche Ativos Atingíveis (agrupados por Tese) + Passivo.

def _achar_aba(wb, *candidatos) -> Optional[str]:
    """Acha o nome real de uma aba tolerando espaços extras/maiúsculas (ex.: 'Fiscal & Cível ')."""
    alvo_norm = [c.strip().casefold() for c in candidatos]
    for nome in wb.sheetnames:
        if nome.strip().casefold() in alvo_norm:
            return nome
    return None


_MAX_LINHAS_VAZIAS_SEGUIDAS = 50  # para de ler apos essa sequencia -- evita varrer um "used
                                  # range" gigante (formatacao ate a ultima linha do Excel)
                                  # so pra confirmar que so tem linha vazia dali pra frente


def _linhas_por_header(ws) -> list:
    """Lê uma aba com cabeçalho na linha 1, devolvendo uma lista de dicts {header: valor}.
    Para de ler cedo se encontrar muitas linhas vazias seguidas (aba com "used range" muito
    maior que os dados reais)."""
    it = ws.iter_rows(values_only=True)
    try:
        header_row = next(it)
    except StopIteration:
        return []
    header = [str(h).strip() if h is not None else "" for h in header_row]
    registros = []
    vazias_seguidas = 0
    for row in it:
        if all(v is None for v in row):
            vazias_seguidas += 1
            if vazias_seguidas >= _MAX_LINHAS_VAZIAS_SEGUIDAS:
                break
            continue
        vazias_seguidas = 0
        registros.append({header[i]: row[i] for i in range(len(header)) if i < len(row)})
    return registros


def _num(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except Exception:
        return None


def ler_matriculas_coleta(wb) -> list:
    # "Detalha imóveis" é a aba de trabalho mais granular (uma linha por item, já com a
    # nota da Tese emendada quando houver) — prioridade sobre "Matrículas", que pode ser
    # uma cópia/versão mais antiga da mesma informação.
    nome = _achar_aba(wb, "Detalha imóveis", "Detalha Imoveis")
    if not nome:
        nome = _achar_aba(wb, "Matrículas", "Matriculas", "Matrículas (2)", "Matriculas (2)")
    return _linhas_por_header(wb[nome]) if nome else []


def ler_processos_coleta(wb, *nomes_aba) -> list:
    nome = _achar_aba(wb, *nomes_aba)
    return _linhas_por_header(wb[nome]) if nome else []


def ler_ecac_coleta(wb) -> list:
    nome = _achar_aba(wb, "E-cac", "e-CAC", "Ecac")
    return _linhas_por_header(wb[nome]) if nome else []


_RE_AREA = re.compile(
    r"([\d]{1,3}(?:\.\d{3})*(?:,\d+)?|\d+(?:,\d+)?)\s*(hectares|ha\b|m[²2])",
    re.IGNORECASE,
)


def _inferir_tipo_e_area(descricao: str) -> Tuple[str, str]:
    """Infere Tipo de Ativo + área a partir da 'Descrição do Imóvel' — heurística por
    palavra-chave, 100% determinística (sem IA, mantendo a arquitetura desse fluxo). É
    melhor esforço: o campo é informativo dentro de um Word revisável manualmente."""
    d = str(descricao or "")
    dl = d.lower()

    area = ""
    m = _RE_AREA.search(d)
    if m:
        unidade = "ha" if m.group(2).lower().startswith(("hectare", "ha")) else "m²"
        area = f"{m.group(1)} {unidade}"

    if "apartamento" in dl:
        tipo = "apartamento"
    elif "casa residencial" in dl or re.search(r"\bcasa\b", dl):
        tipo = "casa residencial"
    elif "sala comercial" in dl:
        tipo = "sala comercial"
    elif any(kw in dl for kw in ("prédio comercial", "predio comercial", "galpão", "galpao",
                                 "loja", "imóvel comercial", "imovel comercial")):
        tipo = "imóvel comercial"
    elif any(kw in dl for kw in ("hectare", "terras rurais", "área rural", "area rural",
                                 "fazenda", "sítio", "sitio")):
        tipo = "área rural"
    elif any(kw in dl for kw in ("terreno", "lote")):
        tipo = "terreno urbano"
    else:
        tipo = ""

    return tipo, area


def _fmt_matricula(matricula) -> str:
    """Formata o número da matrícula com ponto de milhar (ex.: '9596' -> '9.596'). Se já
    vier formatada, com letras, ou não for um número puro, devolve como está."""
    s = str(matricula or "").strip()
    if not s or not s.isdigit():
        return s
    return f"{int(s):,}".replace(",", ".")


_RE_TRANSMISSAO_INICIO = re.compile(r"^(R|AV|Av)\.?\s*[\d]", re.IGNORECASE)


def _ultima_transmissao(celula: str) -> str:
    """Extrai a ÚLTIMA transmissão (entrada 'R.<n>...') da célula de Transmissões —
    confirmado no arquivo real que as entradas vêm em ordem cronológica ascendente,
    separadas por linha em branco OU só por quebra de linha simples, às vezes intercaladas
    com 'AV./Av.' (averbações — não são transmissões de titularidade, são ignoradas)."""
    texto = str(celula or "").strip()
    if not texto:
        return ""
    entradas = []  # [tipo, texto_acumulado]
    for linha in texto.split("\n"):
        l = linha.strip()
        if not l:
            continue
        if _RE_TRANSMISSAO_INICIO.match(l):
            tipo = "R" if l[0].upper() == "R" else "AV"
            entradas.append([tipo, l])
        elif entradas:
            entradas[-1][1] += " " + l
    transmissoes = [t for tipo, t in entradas if tipo == "R"]
    return transmissoes[-1] if transmissoes else texto


def _chave_tese(tese: str) -> str:
    """Normaliza a Tese para fins de agrupamento: usa só a parte antes de uma linha em
    branco. Na prática, às vezes a mesma tese vem com uma nota/explicação emendada na
    mesma célula (ex.: "Fraude à execução - IDPJ (Badi)\n\nElias & CIA transmitiu..."),
    variando de linha pra linha — sem essa normalização, isso vira 2 grupos em vez de 1."""
    primeira_parte = re.split(r"\n\s*\n", tese, maxsplit=1)[0]
    return re.sub(r"\s+", " ", primeira_parte).strip().casefold()


def _agrupar_matriculas_por_tese(linhas: list) -> Tuple[list, list]:
    """Agrupa as linhas da aba Matrículas pela coluna 'Tese' (ordem de 1ª ocorrência),
    tolerando variações de nota emendada na mesma célula (ver _chave_tese). O título
    exibido é a variante mais completa (mais longa) encontrada em cada grupo.
    Devolve (resumo_ativos, ativos_visao_geral) no formato que dossie_ppa.py espera."""
    grupos: dict = {}
    titulos: dict = {}
    ordem: list = []
    for linha in linhas:
        tese_bruta = str(linha.get("Tese") or "").strip() or "Sem tese identificada"
        chave = _chave_tese(tese_bruta)
        if chave not in grupos:
            grupos[chave] = []
            titulos[chave] = tese_bruta
            ordem.append(chave)
        elif len(tese_bruta) > len(titulos[chave]):
            titulos[chave] = tese_bruta
        grupos[chave].append(linha)

    resumo_ativos, ativos_visao_geral = [], []
    for chave in ordem:
        tese = titulos[chave]
        itens = grupos[chave]
        vm_vals    = [v for v in (_num(i.get("Valor da Avaliação Definitiva (VM)")) for i in itens) if v is not None]
        vp_vals    = [v for v in (_num(i.get("Valor da Avaliação Definitiva (VP)")) for i in itens) if v is not None]
        onus_vals  = [v for v in (_num(i.get("Valor Total do Ônus")) for i in itens) if v is not None]

        resumo_ativos.append({
            "tese": tese,
            "vm":   _fmt_valor_br(sum(vm_vals)) if vm_vals else "",
            "vp":   _fmt_valor_br(sum(vp_vals)) if vp_vals else "",
            "onus": _fmt_valor_br(sum(onus_vals)) if onus_vals else "",
            "observacoes": "",
        })

        linhas_visao_geral = []
        for i in itens:
            fracao = i.get("Fração Ideal")
            onus_item = _num(i.get("Valor Total do Ônus"))
            tipo_ativo, area = _inferir_tipo_e_area(i.get("Descrição do Imóvel"))
            linhas_visao_geral.append([
                _fmt_matricula(i.get("Matrícula")),
                tipo_ativo, area, "", "",  # Sit. Produtiva / Liquidez — sem fonte no Excel
                str(fracao) if fracao not in (None, "") else "",
                _fmt_valor_br(onus_item) if onus_item is not None else "Não há",
                "",  # Observações — não transpor (fica em branco a pedido)
            ])
        ativos_visao_geral.append({"tese": tese, "linhas": linhas_visao_geral})

    return resumo_ativos, ativos_visao_geral


def _montar_ativos_detalhados(linhas: list) -> list:
    """Monta a lista 'ativos' no formato que dossie_ppa._filtrar_ativos/
    _preencher_ativos_detalhados esperam, pra preencher as tabelas de imóveis das seções
    narrativas de teses (Penhora Direta/IDPJ/Fraude à Execução) a partir da planilha de
    Coleta de Informações — mantém uma linha por matrícula (não agrupada por tese), já que
    o casamento por palavra-chave roda sobre o texto de Tese de cada item individualmente."""
    ativos = []
    for i in linhas:
        vm = _num(i.get("Valor da Avaliação Definitiva (VM)"))
        vp = _num(i.get("Valor da Avaliação Definitiva (VP)"))
        onus_total = _num(i.get("Valor Total do Ônus"))
        saldo = _num(i.get("Saldo Avaliação - Ônus"))
        ativos.append({
            "tese": str(i.get("Tese") or ""),
            "matricula": _fmt_matricula(i.get("Matrícula")),
            "comarca": str(i.get("Comarca") or ""),
            "proprietario_atual": str(i.get("Proprietário Atual") or ""),
            "descricao": str(i.get("Descrição do Imóvel") or ""),
            "onus_vigentes": str(i.get("Ônus Vigentes") or ""),
            "fracao_atingivel": str(i.get("Fração Ideal") or ""),
            "vm": _fmt_valor_br(vm) if vm is not None else "",
            "vp": _fmt_valor_br(vp) if vp is not None else "",
            "onus_total": _fmt_valor_br(onus_total) if onus_total is not None else "Não há",
            "saldo": _fmt_valor_br(saldo) if saldo is not None else "",
            "transmissoes": _ultima_transmissao(i.get("Transmissões")),
        })
    return ativos


_RE_DOCUMENTO_PARTE = re.compile(r"\s*\((?:CNPJ|CPF)\s*n[ºo°]?\s*[\d./-]+\)", re.IGNORECASE)


def _sem_documento(s: str) -> str:
    """Remove o sufixo '(CNPJ nº ...)'/'(CPF nº ...)' de cada parte, mantendo o '; ' que
    já separa múltiplas partes na célula (não precisamos do CNPJ/CPF no passivo)."""
    return _RE_DOCUMENTO_PARTE.sub("", str(s or "")).strip()


def _chave_coluna(nome) -> str:
    """Normaliza um cabeçalho de coluna: sem acento, sem caixa, sem espaço duplicado."""
    texto = unicodedata.normalize("NFKD", str(nome or ""))
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", texto).strip().casefold()


def _campo(linha: dict, *nomes):
    """Lê uma coluna da planilha da Coleta tolerando acento, caixa e espaço extra no
    cabeçalho ('Polo Passivo ', 'POLO PASSIVO' e 'Polo passivo' são a mesma coluna).
    Aceita sinônimos na ordem informada e devolve o primeiro valor não vazio."""
    normalizados = {_chave_coluna(k): v for k, v in linha.items()}
    for nome in nomes:
        valor = normalizados.get(_chave_coluna(nome))
        if valor not in (None, ""):
            return valor
    return None


def _linha_processo_coleta(linha: dict) -> dict:
    # "Executado" do dossiê sai do POLO PASSIVO da planilha (quem é réu/executado no
    # processo), não de "Vinculado à" — essa coluna diz apenas a quem a pesquisa estava
    # atrelada, e frequentemente não é a parte que figura no polo passivo.
    return {
        "cnj":     str(_campo(linha, "Número CNJ", "N° CNJ", "Nº CNJ") or ""),
        "vinc":    _sem_documento(_campo(linha, "Vinculado à", "Vinculado a") or ""),
        "passivo": _sem_documento(_campo(linha, "Polo passivo", "Partes passivas", "Polo Passivo") or ""),
        "data":   _parse_date(str(_campo(linha, "Data da distribuição", "Distribuição") or "")) or "",
        "ativo":  _sem_documento(_campo(linha, "Polo ativo", "Partes ativas") or ""),
        "valor":  _num(_campo(linha, "Valor da causa")),
        "status": str(_campo(linha, "Situação atual", "Situação", "Status") or ""),
        "sat":    _num(_campo(linha, "Saldo Atualizado estimado", "SAT estimado")),
    }


def _arquivado_definitivamente(status: str) -> bool:
    """Detecta 'arquivado/arquivamento definitivo' por palavra-chave (não por igualdade
    exata) — a planilha real usa variações como 'Arquivamento Definitivo', não só
    'Arquivado'/'Arquivado Definitivamente'. Arquivamento PROVISÓRIO ou um 'Arquivamento'
    sem qualificador NÃO contam como definitivo — só excluímos quando "definitiv" aparece
    explicitamente junto de "arquiv"."""
    s = str(status or "").upper()
    return "ARQUIV" in s and "DEFINITIV" in s


def _excluir_arquivados(processos: list) -> list:
    """Passivo: nunca lista processos já arquivados definitivamente."""
    return [p for p in processos if not _arquivado_definitivamente(p.get("status", ""))]


def _classificar_fiscal_civel_coleta(linhas: list) -> Tuple[list, list]:
    """A aba 'Fiscal & Cível' mistura os dois — separa pela coluna Classe (mesmo
    critério de _is_fiscal já usado no fluxo da Predictus)."""
    fiscais, civeis = [], []
    for linha in linhas:
        classe = str(linha.get("Classe") or "")
        proc = _linha_processo_coleta(linha)
        (fiscais if _is_fiscal("", classe) else civeis).append(proc)
    return _excluir_arquivados(fiscais), _excluir_arquivados(civeis)


def _ler_ecac_processado(linhas: list) -> list:
    resultado = []
    for linha in linhas:
        resultado.append({
            "nome":     str(linha.get("Nome") or "").strip(),
            "cpf_cnpj": str(linha.get("CPF/CNPJ") or "").strip(),
            "saldo":    _num(linha.get("E-cac")),
        })
    return resultado


def coleta_gerar_dossie_ativos_passivo(excel_coleta_file, dossie_file) -> Tuple[str, str, Any]:
    """
    Preenche Ativos Atingíveis (Visão Consolidada + Visão Geral, agrupados por Tese) e
    Passivo (Fiscal/Trabalhista/Cível/e-CAC) de um dossiê PPA a partir do Excel de
    "Coleta de Informações sobre Caso" (abas Matrículas / Fiscal & Cível / Trabalhista / E-cac).
    Retorna: (log_text, status_msg, caminho_do_docx_ou_None).
    """
    lines: list = []

    def log(msg: str):
        lines.append(msg)
        return "\n".join(lines)

    if not excel_coleta_file:
        return "Nenhum Excel da Coleta de Informações enviado.", "Erro: nenhum Excel enviado", None

    fp = excel_coleta_file.name if hasattr(excel_coleta_file, "name") else str(excel_coleta_file)
    log(f"📂 {os.path.basename(fp)}")

    try:
        # read_only=True evita carregar estilos/dimensoes de abas com "used range" gigante
        # (ex.: formatacao aplicada ate a ultima linha do Excel) -- sem isso, o carregamento
        # de planilhas assim pode travar por minutos mesmo com poucas linhas de dado real.
        wb = load_workbook(fp, data_only=True, read_only=True)
    except Exception as exc:
        return log(f"\n❌ Erro ao abrir o Excel: {exc}"), "Erro ao abrir o Excel", None

    linhas_matriculas    = ler_matriculas_coleta(wb)
    linhas_fiscal_civel  = ler_processos_coleta(wb, "Fiscal & Cível", "Fiscal & Civel")
    linhas_trabalhista   = ler_processos_coleta(wb, "Trabalhista")
    linhas_ecac          = ler_ecac_coleta(wb)

    if not linhas_matriculas and not linhas_fiscal_civel and not linhas_trabalhista and not linhas_ecac:
        return log(
            "\n⚠️ Nenhuma das abas esperadas (Matrículas / Fiscal & Cível / Trabalhista / E-cac) "
            "foi encontrada ou está vazia."
        ), "Nenhum dado encontrado no Excel", None

    resumo_ativos, ativos_visao_geral = _agrupar_matriculas_por_tese(linhas_matriculas)
    ativos_detalhados = _montar_ativos_detalhados(linhas_matriculas)
    fiscais, civeis = _classificar_fiscal_civel_coleta(linhas_fiscal_civel)
    trabalhistas = _excluir_arquivados([_linha_processo_coleta(l) for l in linhas_trabalhista])
    ecac = _ler_ecac_processado(linhas_ecac)

    log(f"   🏠 {len(linhas_matriculas)} matrícula(s) em {len(resumo_ativos)} tese(s)")
    log(f"   ⚖️  {len(fiscais)} fiscal | {len(trabalhistas)} trabalhista | {len(civeis)} cível")
    log(f"   💰 {len(ecac)} registro(s) de e-CAC")

    dossie_path = None
    if dossie_file:
        dossie_path = dossie_file.name if hasattr(dossie_file, "name") else str(dossie_file)
        log(f"\n📄 Atualizando dossiê enviado: {os.path.basename(dossie_path)}")
    else:
        log("\n📄 Nenhum dossiê enviado — gerando esqueleto com ativos/passivo preenchidos.")

    try:
        out = preencher_ativos_e_passivo_dossie(
            dossie_path, resumo_ativos, ativos_visao_geral, fiscais, trabalhistas, civeis, ecac,
            ativos_detalhados,
        )
    except Exception as exc:
        return log(f"\n❌ Erro ao preencher o dossiê: {exc}"), "Erro ao preencher o dossiê", None

    log(f"\n✅ Ativos ({len(resumo_ativos)} tese(s)) e passivo preenchidos.")
    status = f"{len(resumo_ativos)} tese(s) de ativos + passivo preenchidos"
    return "\n".join(lines), status, out


def coleta_gerar_dossie_dispatch(excel_predictus_files, excel_coleta_file, dossie_file) -> Tuple[str, str, Any]:
    """Escolhe o fluxo de preenchimento do dossiê: se o Excel da Coleta de Informações
    (Matrículas/Fiscal & Cível/Trabalhista/E-cac) foi enviado, usa o fluxo novo (Ativos +
    Passivo); senão, mantém o fluxo antigo (só Passivo, via Excel da Predictus)."""
    if excel_coleta_file:
        return coleta_gerar_dossie_ativos_passivo(excel_coleta_file, dossie_file)
    return coleta_gerar_dossie(excel_predictus_files, dossie_file)
