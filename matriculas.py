# -*- coding: utf-8 -*-
import os
import re
import math
import unicodedata
import time
import tempfile
import concurrent.futures
from datetime import date
from pathlib import Path
from typing import Optional, List

from dateutil.relativedelta import relativedelta

import pandas as pd
from pydantic import BaseModel
from google.genai import types

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from utils import (
    _get_gemini_clients,
    _erro_gemini_permite_failover,
    _paginas_digitalizadas_pdf,
    _retry,
    _responder_pergunta_generica,
    GEMINI_MODEL_EXTRACAO,
    GEMINI_MODEL_OCR,
    GEMINI_MODEL_RELATORIO,
    GEMINI_MODEL_QA,
)
from legal_prompts import REGRA_REFERENCIA_MATRICULA


try:
    MATRICULAS_MAX_WORKERS = max(1, int(os.getenv("MATRICULAS_MAX_WORKERS", "3")))
except ValueError:
    MATRICULAS_MAX_WORKERS = 3


def _mat_get_clients():
    return _get_gemini_clients()


# ── Modelos Pydantic ──────────────────────────────────────────────────────────

class OnusFinanceiro(BaseModel):
    codigo: Optional[str] = None
    tipo: Optional[str] = None          # penhora, hipoteca, alienacao fiduciaria, CCB, etc.
    credor: Optional[str] = None        # credor/beneficiario da garantia (banco/instituicao)
    numero_processo: Optional[str] = None   # nº do processo judicial (penhora, arresto, execucao)
    numero_instrumento: Optional[str] = None  # nº da CCB, contrato, operacao bancaria
    valor_principal: Optional[float] = None
    moeda: Optional[str] = None       # BRL/Real ou a moeda histórica indicada no ato
    data_celebracao: Optional[str] = None
    parcela_mensal: Optional[float] = None
    vencimento_final: Optional[str] = None
    cancelado: Optional[bool] = False   # True se o ato foi baixado/cancelado por AV/R posterior


class Transmissao(BaseModel):
    codigo: Optional[str] = None        # ex: "R.4"
    tipo: Optional[str] = None          # Compra e Venda, Doação, Integralização, etc.
    data: Optional[str] = None          # DD/MM/AAAA
    de_nome: Optional[str] = None
    para_nome: Optional[str] = None
    de_doc: Optional[str] = None        # CPF/CNPJ sem formatacao (so digitos)
    para_doc: Optional[str] = None


class MatriculaExtraida(BaseModel):
    matricula: Optional[str] = None
    comarca: Optional[str] = None
    proprietario_atual: Optional[str] = None
    fracao_ideal: Optional[str] = None
    descricao_imovel: Optional[str] = None
    transmissoes_averbadas_registradas: Optional[str] = None
    onus_vigentes_registrados_averbados: Optional[str] = None
    observacoes: Optional[str] = None
    onus_cancelados: Optional[str] = None
    grau_confianca: Optional[str] = None
    onus_financeiros: Optional[List[OnusFinanceiro]] = None
    transmissoes_estruturadas: Optional[List[Transmissao]] = None


# ── Prompt ────────────────────────────────────────────────────────────────────

PROMPT_EXTRACAO_MATRICULA = (
    "Leia integralmente esta matrícula imobiliária brasileira, inclusive texto manuscrito, carimbos, "
    "continuações e páginas digitalizadas. Produza uma extração factual completa para uma segunda etapa "
    "de consolidação jurídica. Não resuma e não conclua por amostragem.\n\n"
    "Faça obrigatoriamente:\n"
    "1. Identifique número, cartório, comarca e eventual encerramento, baixa, rematriculação ou matrículas "
    "originadas/derivadas.\n"
    "2. Transcreva em ordem todos os registros R e averbações AV, sem omitir nenhum código, indicando "
    "data, natureza, partes, CPF/CNPJ, frações, valores, moeda e referências a atos posteriores.\n"
    "3. Reconstrua a cadeia dominial. Destaque a última transmissão válida — inclusive compra e venda, "
    "arrematação, adjudicação, dação em pagamento, doação, partilha, integralização ou sucessão — e as "
    "frações que permaneceram com cada proprietário após todos os atos.\n"
    "4. Monte um inventário de todos os ônus e restrições. Para cada um, diga se está vigente ou qual ato "
    "o cancelou/baixou/levantou. Nenhum ônus pode ficar sem uma dessas duas classificações.\n"
    "5. Preserve literalmente moedas históricas e não converta valores. Se o ônus em moeda antiga não foi "
    "cancelado, classifique-o como vigente e sinalize que o valor não é calculável automaticamente.\n"
    "6. Não use Markdown, negrito ou pares de asteriscos. Nunca invente informação ilegível ou ausente.\n\n"
    + REGRA_REFERENCIA_MATRICULA
)


PROMPT_MATRICULA = (
    "Voce e um assistente especializado em analise de matriculas de imoveis brasileiros.\n\n"
    "Consolide a extracao factual integral fornecida abaixo em formato estruturado. Resolva a cadeia "
    "dominial e a situacao de cada onus antes de preencher o JSON.\n\n"
    "Preencha os seguintes campos:\n"
    "- matricula\n- comarca\n- proprietario_atual\n- fracao_ideal\n- descricao_imovel\n"
    "- transmissoes_averbadas_registradas\n- onus_vigentes_registrados_averbados\n"
    "- observacoes\n- onus_cancelados\n- grau_confianca\n"
    "- onus_financeiros\n- transmissoes_estruturadas\n\n"
    "Regras obrigatorias:\n\n"
    "1. matricula: comece pelo numero e, no MESMO campo, informe imediatamente se houve baixa, "
    "encerramento, rematriculacao ou origem de outra matricula. Exemplo: 21.101 - Encerrada, deu origem "
    "à matrícula 73.923. Nao deixe essa informacao apenas em observacoes.\n\n"
    "1.1. proprietario_atual: reconstrua toda a cadeia dominial e use o adquirente da ULTIMA transmissao "
    "valida, qualquer que seja sua natureza: compra e venda, arrematacao, adjudicacao, dacao em pagamento, "
    "doacao, partilha, integralizacao, sucessao ou outra transferencia de dominio. Um proprietario anterior "
    "nao pode prevalecer sobre transmissao posterior. Inclua CPF/CNPJ quando constar, sempre entre parenteses.\n"
    "   Se houver copropriedade, some aquisicoes e alienacoes parciais e indique a proporcao final de CADA "
    "titular. Exemplo: Fulano da Silva (CPF 123.456.789-00) - 40%; Beltrano Camargo (CPF 987.654.321-00) - 60%.\n"
    "   fracao_ideal: repita de forma objetiva o quadro final, no formato Nome: percentual ou fracao. "
    "Confira se as proporcoes totalizam 100% ou a integralidade registral; se nao for possivel fechar, "
    "informe a divergencia sem inventar.\n\n"
    "2. comarca: indicar a comarca e qual é o cartório. Exemplo: '1º CRI de Curitiba/PR'.\n\n"
    "3. descricao_imovel: resumo em ate 3 frases. Tipo, localizacao, area, identificadores.\n\n"
    "4. transmissoes_averbadas_registradas: todas, cronologicas, sinteticas.\n"
    "   Inclua o CPF/CNPJ de cada pessoa, mas APENAS NA PRIMEIRA VEZ que ela aparecer nesta celula.\n"
    "   Diga sempre o preço do imóvel em caso de compra e venda ou integralizacao.\n"
    "   Inclua registros anteriores a abertura se disponiveis.\n"
    "   Exemplo:\n"
    "   R.4 - Compra e Venda: Jose Marques da Cruz (CPF 123.456.789-00) vendeu para Carlos Guimaraes"
    " (CPF 987.654.321-00) em 15/12/1986, pelo valor de R$ 2.000.000,00.\n\n"
    "   R.5 - Doação: Carlos Guimaraes doou para Murilo Menezes (CPF 111.222.333-44) em 03/06/1988.\n\n"
    "   Nao use negrito, Markdown ou asteriscos na natureza das transmissoes.\n\n"
    "5. onus_vigentes_registrados_averbados: inclua TODOS os onus, gravames e restricoes ainda vigentes, "
    "financeiros ou nao, incluindo penhora, arresto, hipoteca, alienacao fiduciaria, indisponibilidade, "
    "averbacao premonitoria, averbacao de ajuizamento/execucao, protesto e restricao administrativa.\n"
    "   Se nao houver nenhum onus vigente, preencha exatamente: Sem onus vigentes identificados.\n\n"
    "6. Formato de cada onus vigente:\n"
    "   [Codigo]: [Tipo], [n. cedula se disponivel], [dd/mm/aaaa], [partes com CPF/CNPJ entre parenteses],\n"
    "   vencimento [data se disponivel], Valor principal: R$ [valor se disponivel].\n"
    "   OBRIGATORIO para penhora, arresto, execucao ou ajuizamento: inclua o numero do processo.\n"
    "   Exemplo: AV.12: Penhora, 12/03/2023, Exequente: Banco XYZ; Executado: Joao Silva,"
    " Proc. 1234567-89.2023.8.26.0001, Valor principal: R$ 250.000,00.\n"
    "   Vara e comarca vao em observacoes, nao aqui.\n\n"
    "7. Todo onus identificado deve aparecer em EXATAMENTE um dos campos: onus_vigentes ou onus_cancelados. "
    "Se nao estiver vigente, deve obrigatoriamente constar em onus_cancelados, com o ato de baixa quando "
    "identificavel. Nunca deixe um onus somente em observacoes.\n\n"
    "8. PDF escaneado: extraia o maximo possivel do conteudo visual.\n\n"
    "9. Nunca invente dados. Nomes: primeira letra maiuscula, exceto artigos (de, da, do, das, dos, e).\n"
    "   Erros de grafia: copie como esta e registre em observacoes.\n\n"
    "10. observacoes: inclua SEMPRE que houver:\n"
    "   - Averbacoes relevantes (divorcio, matrimonio, georreferenciamento, reserva legal, quitacao)\n"
    "   - Irregularidades, CPFs divergentes para a mesma pessoa, erros de grafia\n"
    "   - Matrícula encerrada, rematricula, baixa legibilidade\n"
    "   Nao use observacoes como destino unico de qualquer onus: onus vigentes e cancelados pertencem aos "
    "campos proprios. Informacoes complementares podem ser repetidas aqui apenas quando indispensaveis.\n"
    "   Se nao houver: Sem observacoes relevantes. Nunca deixe vazio.\n\n"
    "11. onus_cancelados: TODOS os onus cancelados, baixados ou levantados, financeiros ou nao, sempre "
    "indicando qual ato cancelou quando essa referencia estiver disponivel.\n"
    "    Exemplo: AV.9: Penhora cancelada pela AV.22.\n"
    "    NAO repita esses onus em observacoes.\n"
    "    Se nao houver: Sem onus cancelados identificados.\n\n"
    "12. grau_confianca: Alto, Medio ou Baixo.\n"
    "    Medio/Baixo: justificativa entre parenteses. Ex: Medio (Baixa legibilidade da AV-3).\n"
    "    Alto: preencha apenas Alto.\n"
    "    A marca d'agua 'nao possui valor de certidao' por si so nao e indicio de confianca reduzida.\n"
    "    ATENCAO: matricula encerrada ou rematriculada = obrigatoriamente Medio ou Baixo.\n\n"
    "13. Responda apenas com JSON valido aderente ao schema.\n\n"
    "14. Redacao padronizada e objetiva. Evite textos longos.\n\n"
    "15. Pule UMA LINHA entre itens autonomos diferentes em cada campo de texto.\n\n"
    "16. NUNCA pule linha dentro de uma mesma frase. Intervalos como AV.1 a AV.8 ou"
    " referencias como conforme R.5 sao parte da frase e NAO devem ter quebra de linha.\n\n"
    "17. Apos a analise, verifique todos os AVs/Rs para ver se algum ficou faltando.\n"
    "    Caso falte algum, indique em observacoes qual registro ou averbaçao nao pode ser identificado.\n\n"
    "18. Para cada onus de natureza financeira encontrado (penhora, hipoteca, alienacao fiduciaria,"
    " CCB, confissao de divida ou qualquer titulo de credito), preencha onus_financeiros com:\n"
    "   - codigo: codigo do ato (ex: R.5, AV.9)\n"
    "   - tipo: tipo do onus (ex: penhora, hipoteca, alienacao fiduciaria, CCB)\n"
    "   - credor: OBRIGATORIO para hipoteca e alienacao fiduciaria. Nome do credor/beneficiario"
    " da garantia (banco ou instituicao financeira). Apenas o nome, sem CPF/CNPJ."
    " Null se nao identificavel ou se o tipo de onus nao for hipoteca/alienacao fiduciaria.\n"
    "   - numero_processo: OBRIGATORIO para penhora, arresto, execucao e qualquer ato judicial."
    " Informe o numero completo do processo (ex: 0001234-56.2023.8.26.0100). Null se nao for ato judicial.\n"
    "   - numero_instrumento: OBRIGATORIO para CCB, contrato bancario, operacao de credito, cedula."
    " Informe o numero da CCB, do contrato ou da operacao (ex: CCB 12345, Contrato 67890, Op. 00123)."
    " Null se nao for instrumento de credito.\n"
    "   - valor_principal: numero puro na moeda indicada, sem simbolo ou separador de milhar\n"
    "   - moeda: BRL/Real para valores em reais; para moeda historica, informe literalmente Cruzeiro, "
    "Cruzado, NCz$, Cr$, URV ou a denominacao constante do ato.\n"
    "   - data_celebracao: data do ato em DD/MM/AAAA\n"
    "   - parcela_mensal: valor da parcela mensal se explicitamente indicada. Null se nao houver.\n"
    "   - vencimento_final: data de vencimento final em DD/MM/AAAA. Null se nao houver.\n"
    "   - cancelado: true se houver AV/R posterior que baixou, cancelou ou levantou este onus."
    " false se ainda vigente.\n"
    "   NAO inclua: indisponibilidade, averbacao premonitoria, averbacao de ajuizamento,"
    " averbacao de execucao, protesto.\n"
    "   Averbacao de execucao e averbacao premonitoria sao NOTACOES PROCESSUAIS, nao gravames financeiros"
    " — ignorar no calculo mesmo que tenham valor indicado.\n"
    "   Inclua tambem onus financeiros sem valor identificavel e use valor_principal=null. Nao invente dados.\n\n"
    "19. VARREDURA OBRIGATORIA AV/R POR AV/R:\n"
    "    a) Identifique a numeracao maxima de R e de AV no documento.\n"
    "    b) Percorra individualmente de R.1 ate o ultimo R, e de AV.1 ate o ultimo AV.\n"
    "    c) Para CADA item, classifique em EXATAMENTE uma destas categorias:\n"
    "       - ONUS VIGENTE: todo gravame ou restricao ainda ativo, financeiro ou nao -> onus_vigentes\n"
    "       - ONUS CANCELADO: onus cuja AV/R posterior cancelou/baixou -> onus_cancelados\n"
    "       - OBSERVACAO/TRANSMISSAO: compra e venda, doacao, partilha, integralizacao, divorcio,"
    " retificacao, georreferenciamento, encerramento, etc.\n"
    "    d) NUNCA pule nenhum R/AV. Se ilegivel, cite em observacoes. Omissao e ERRO GRAVE.\n"
    "    e) Antes de finalizar, confira se TODOS os R/AV aparecem em alguma categoria.\n\n"
    "20. transmissoes_estruturadas: para CADA transmissao (compra e venda, doacao, partilha,"
    " integralizacao, etc.) identificada, preencha uma entrada com:\n"
    "   - codigo: codigo do ato (ex: 'R.4')\n"
    "   - tipo: tipo da transmissao (ex: 'Compra e Venda', 'Doacao', 'Integralizacao')\n"
    "   - data: data do ato em DD/MM/AAAA. Null se nao identificada.\n"
    "   - de_nome: nome do transmitente (quem vendeu/doou/transferiu)\n"
    "   - para_nome: nome do adquirente (quem comprou/recebeu)\n"
    "   - de_doc: CPF ou CNPJ do transmitente, APENAS DIGITOS sem formatacao (ex: '12345678901')."
    " Null se nao disponivel.\n"
    "   - para_doc: CPF ou CNPJ do adquirente, APENAS DIGITOS. Null se nao disponivel.\n"
    "   Nao invente. Deixe null qualquer campo nao identificavel.\n\n"
    "21. CONSISTENCIA OBRIGATORIA DOS ONUS:\n"
    "   Faça uma lista de controle de todo R/AV que constitua onus ou restricao. Cada item deve aparecer "
    "em exatamente um campo textual: onus_vigentes_registrados_averbados OU onus_cancelados.\n"
    "   Para cada item em onus_financeiros com cancelado=false, confirme sua presenca em onus_vigentes. "
    "Para cada item com cancelado=true, confirme sua presenca em onus_cancelados.\n"
    "   A ausencia em ambos os campos ou a presenca simultanea nos dois e ERRO GRAVE.\n\n"
    "22. VALORES EM MOEDA PRE-REAL (anterior a julho/1994): qualquer onus cujo valor esteja expresso\n"
    "   em Cruzeiros, Cruzados, NCz$, Cr$, URV ou qualquer moeda anterior ao Real continua sendo onus "
    "vigente se nao houver baixa/cancelamento. Inclua-o em onus_vigentes e em onus_financeiros, preserve "
    "a moeda original no campo moeda e escreva no texto: valor nao calculado automaticamente por estar "
    "expresso em moeda historica. Nao converta nem some esse valor ao total.\n"
    "   Se nao houver informacao sobre a moeda mas o valor parecer incompativel com imóveis brasileiros\n"
    "   (ex: valores acima de R$ 500.000.000 para credito rural dos anos 80-90), trate como moeda antiga.\n\n"
    "23. FORMATACAO FINAL: nunca use Markdown, negrito ou pares de asteriscos. Sempre que CPF ou CNPJ "
    "aparecer em qualquer campo textual, mantenha o documento entre parenteses, por exemplo: Nome "
    "da Pessoa (CPF 123.456.789-00) ou Empresa S/A (CNPJ 12.345.678/0001-90).\n\n"
    + REGRA_REFERENCIA_MATRICULA
)


# ── Helpers de texto ──────────────────────────────────────────────────────────

MINUSC_SET = {"de", "da", "do", "das", "dos", "e", "a", "o", "em", "com", "por", "ao", "aos", "as"}
_ITEM_RE = re.compile(
    r"(?<!\n\n)(?<!\n)(?=(?:R|AV)-\d+[-–]\d[\d.]*\s*(?:\([^)]*\)\s*)?:)",
    re.IGNORECASE,
)


def _mat_capitalizar_token(p):
    core = p.strip(".,;:()")
    inicio = p.find(core) if core else 0
    prefix = p[:inicio]
    suffix = p[inicio + len(core):]
    if not core: return p
    low = core.lower()
    if low in MINUSC_SET: return prefix + low + suffix
    if "/" in core: return prefix + core.upper() + suffix
    if re.match(r"^[A-Z]{2,6}$", core): return prefix + core + suffix
    return prefix + core.capitalize() + suffix


def _mat_corrigir_caps(texto):
    if not texto or pd.isna(texto): return texto
    resultado = []
    for linha in str(texto).split("\n"):
        nova = []
        for p in linha.split(" "):
            core = p.strip(".,;:()")
            if core.isupper() and len(core) > 2 and "/" not in core and not re.match(r"^\d", core):
                nova.append(_mat_capitalizar_token(p))
            else:
                nova.append(p)
        resultado.append(" ".join(nova))
    return "\n".join(resultado)


_DOC_ROTULADO_RE = re.compile(
    r"\b(?P<tipo>CPF(?:/MF)?|CNPJ)\s*(?:n[º°o.]?\s*)?"
    r"(?P<doc>(?:\d{2}\.?\d{3}\.?\d{3}/?\d{4}-?\d{2}|\d{3}\.?\d{3}\.?\d{3}-?\d{2}))",
    flags=re.IGNORECASE,
)


def _mat_formatar_documento(doc: str) -> str:
    digitos = re.sub(r"\D", "", doc or "")
    if len(digitos) == 11:
        return f"{digitos[:3]}.{digitos[3:6]}.{digitos[6:9]}-{digitos[9:]}"
    if len(digitos) == 14:
        return f"{digitos[:2]}.{digitos[2:5]}.{digitos[5:8]}/{digitos[8:12]}-{digitos[12:]}"
    return re.sub(r"\s+", "", doc or "")


def _mat_parentetizar_documentos(texto):
    """Mantém CPF/CNPJ visíveis e garante que cada documento esteja entre parênteses."""
    if not texto or pd.isna(texto):
        return texto
    original = str(texto)

    def _sub(match):
        tipo = "CNPJ" if match.group("tipo").upper() == "CNPJ" else "CPF"
        rotulo = f"{tipo} {_mat_formatar_documento(match.group('doc'))}"
        antes = original[:match.start()].rstrip()
        depois = original[match.end():].lstrip()
        if antes.endswith("(") and depois.startswith(")"):
            return rotulo
        return f"({rotulo})"

    return _DOC_ROTULADO_RE.sub(_sub, original)


def _mat_normalizar_quebras(texto):
    if not texto or pd.isna(texto): return texto
    t = str(texto).replace("\r\n", "\n").replace("\r", "\n")
    t = _ITEM_RE.sub("\n\n", t)
    t = re.sub(r"\n{3,}", "\n\n", t)
    return t.strip()


def _mat_normalizar_siglas(texto):
    if not texto or pd.isna(texto): return texto
    t = str(texto)
    t = re.sub(r"\bS/[Aa]\b", "S/A", t)
    t = re.sub(r"\bccir\b", "CCIR", t, flags=re.IGNORECASE)
    t = re.sub(r"\bcpf(?:/mf)?\b", "CPF", t, flags=re.IGNORECASE)
    t = re.sub(r"\bcnpj\b", "CNPJ", t, flags=re.IGNORECASE)
    return t


def _fmt_brl(valor: float) -> str:
    return f"R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


# ── Pós-processamento ─────────────────────────────────────────────────────────

def _mat_pos_processar(resultado):
    # Todos os campos de texto: sem marcadores Markdown, com documentos entre
    # parênteses e siglas normalizadas.
    campos_texto = ["matricula", "comarca", "proprietario_atual", "fracao_ideal",
                    "descricao_imovel",
                    "transmissoes_averbadas_registradas",
                    "onus_vigentes_registrados_averbados", "observacoes", "onus_cancelados",
                    "grau_confianca"]
    for campo in campos_texto:
        v = resultado.get(campo)
        if v:
            v = str(v).replace("**", "")
            v = _mat_parentetizar_documentos(v)
            v = _mat_corrigir_caps(v)
            v = _mat_normalizar_siglas(v)
            resultado[campo] = v

    # Normalizar quebras de linha nos campos longos
    for campo in ["transmissoes_averbadas_registradas", "onus_vigentes_registrados_averbados",
                  "observacoes", "onus_cancelados"]:
        v = resultado.get(campo)
        if v:
            resultado[campo] = _mat_normalizar_quebras(_mat_normalizar_siglas(v))

    return resultado


# ── Cálculo de ônus ───────────────────────────────────────────────────────────

_TIPOS_NAO_FINANCEIROS = {
    "indisponibilidade", "premonitoria", "premonitório", "premonotoria",
    "ajuizamento", "protesto", "restricao administrativa",
    "averbacao de execucao", "averbação de execução",
    "averbacao execucao", "averbação execução",
}


def _mat_calcular_valor_onus(onus_list: list, onus_vigentes_texto: str = "",
                              data_referencia: date = None) -> str:
    if not onus_list:
        return ""
    # Se o campo textual já diz "sem ônus vigentes", o modelo concluiu que
    # não há gravames ativos — não computar valor (evita valores em moeda antiga
    # ou itens que o modelo identificou como sem efeito prático serem somados).
    if onus_vigentes_texto and "sem onus vigentes" in onus_vigentes_texto.lower():
        return ""
    if data_referencia is None:
        data_referencia = date.today()

    linhas_individuais = []
    total = 0.0

    for onus in onus_list:
        def _get(key):
            return onus.get(key) if isinstance(onus, dict) else getattr(onus, key, None)

        cancelado = _get("cancelado")
        if cancelado:
            continue

        codigo             = _get("codigo")
        tipo               = _get("tipo") or ""
        numero_processo    = _get("numero_processo")
        numero_instrumento = _get("numero_instrumento")
        valor              = _get("valor_principal")
        moeda              = (_get("moeda") or "BRL").strip()
        data_str           = _get("data_celebracao")
        parcela            = _get("parcela_mensal")

        # Ignora tipos não financeiros (indisponibilidade, premonitória, etc.)
        tipo_lower = tipo.lower()
        if any(t in tipo_lower for t in _TIPOS_NAO_FINANCEIROS):
            continue

        moeda_norm = _mat_normalizar_texto(moeda)
        if moeda_norm not in {"brl", "real", "reais", "r$"}:
            # O ônus continua vigente no campo textual, mas valores históricos
            # não são convertidos nem somados automaticamente.
            continue

        if not valor or not data_str:
            continue
        try:
            partes = data_str.strip().split("/")
            data_onus = date(int(partes[2]), int(partes[1]), int(partes[0]))
        except Exception:
            continue

        delta = relativedelta(data_referencia, data_onus)
        meses = delta.years * 12 + delta.months
        if meses < 0:
            continue

        valor_atualizado = valor * (1 + 0.01 * meses)
        total += valor_atualizado

        prefixo = f"∘ {codigo}" if codigo else "∘"
        if tipo:
            prefixo += f" ({tipo})"
        prefixo += ": "

        linha = prefixo + _fmt_brl(valor_atualizado)

        # Número do processo (penhoras, arrestos, execuções)
        if numero_processo:
            linha += f"\n   Proc. {numero_processo}"

        # Número do instrumento (CCB, contrato, operação)
        if numero_instrumento:
            linha += f"\n   {numero_instrumento}"

        if parcela and parcela > 0:
            saldo = valor_atualizado - parcela * meses
            if saldo > 0:
                linha += f"\n   (em caso de adimplência, saldo de {_fmt_brl(saldo)})"

        linhas_individuais.append(linha)

    if not linhas_individuais:
        return ""

    return f"Total: {_fmt_brl(total)}\n" + "\n".join(linhas_individuais)


def _mat_normalizar_texto(s: str) -> str:
    if not s:
        return ""
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return s.lower().strip()


def _mat_resumo_garantias(onus_list: list) -> str:
    """Resumo enxuto das garantias reais vigentes (Alienação Fiduciária e Hipoteca),
    listando apenas os credores — sem grau, valor ou demais detalhes."""
    af_credores, hip_credores = [], []
    af_vistos, hip_vistos = set(), set()

    for onus in onus_list or []:
        def _get(key):
            return onus.get(key) if isinstance(onus, dict) else getattr(onus, key, None)

        if _get("cancelado"):
            continue

        tipo_norm = _mat_normalizar_texto(_get("tipo") or "")
        if "alienacao fiduciaria" in tipo_norm:
            destino, vistos = af_credores, af_vistos
        elif "hipoteca" in tipo_norm:
            destino, vistos = hip_credores, hip_vistos
        else:
            continue

        credor = (_get("credor") or "").strip() or "credor não identificado"
        chave = _mat_normalizar_texto(credor)
        if chave in vistos:
            continue
        vistos.add(chave)
        destino.append(credor)

    partes = []
    if af_credores:
        partes.append("Alienação Fiduciária: " + ", ".join(af_credores))
    if hip_credores:
        partes.append("Hipoteca: " + ", ".join(hip_credores))

    return "\n".join(partes) if partes else "Sem garantia real vigente identificada."


# ── Parsing de devedores/relacionados ─────────────────────────────────────────

# Nome curto casa com qualquer coisa por conter-se em outro nome ("Ana" dentro de
# "Ana Paula"). Abaixo deste tamanho, só igualdade exata conta.
_MIN_NOME_PARA_CONTER = 8


def _mat_chave_nome(nome) -> str:
    """Normaliza um nome para comparação: sem acento, sem pontuação, sem caixa."""
    texto = unicodedata.normalize("NFKD", str(nome or ""))
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    texto = re.sub(r"[^\w\s]", " ", texto)
    return re.sub(r"\s+", " ", texto).strip().casefold()


def _mat_nomes_casam(chave_a: str, chave_b: str) -> bool:
    """Igualdade, ou um nome contido no outro como sequência inteira de palavras.

    A conteção nos dois sentidos cobre o caso real: a matrícula traz "Agropecuária
    Teste Ltda - EPP" e a usuária digitou "Agropecuária Teste Ltda", ou o contrário.
    """
    if not chave_a or not chave_b:
        return False
    if chave_a == chave_b:
        return True
    menor, maior = sorted((chave_a, chave_b), key=len)
    if len(menor) < _MIN_NOME_PARA_CONTER:
        return False
    return re.search(rf"(?:^|\s){re.escape(menor)}(?:\s|$)", maior) is not None


def _mat_parse_parties(pares) -> list:
    """Normaliza os pares (nome, CPF/CNPJ) vindos da interface.

    Antes esta função lia texto livre e devolvia SÓ um set de documentos — quem
    digitasse apenas o nome não gerava alerta nenhum, em silêncio. Agora o nome é
    preservado e vira critério de busca por si.
    """
    partes = []
    for nome, doc in pares or []:
        doc_norm = re.sub(r"[^\d]", "", str(doc or ""))
        if len(doc_norm) not in (11, 14):  # nem CPF nem CNPJ
            doc_norm = ""
        chave = _mat_chave_nome(nome)
        if not doc_norm and not chave:
            continue
        partes.append({"nome": str(nome or "").strip(), "doc": doc_norm, "chave": chave})
    return partes


def _mat_parte_envolvida(partes: list, doc: str, nome: str) -> bool:
    """Casa por documento quando ele existe; senão, cai para o nome."""
    doc_norm = re.sub(r"[^\d]", "", str(doc or ""))
    chave = _mat_chave_nome(nome)
    for parte in partes:
        if parte["doc"] and doc_norm and parte["doc"] == doc_norm:
            return True
        if _mat_nomes_casam(parte["chave"], chave):
            return True
    return False


def _mat_detectar_alertas(resultado: dict, devedores: list,
                           relacionados: list, data_ajuizamento: date | None) -> set:
    """
    Retorna set com alertas detectados: 'amarelo' e/ou 'vermelho'.
    - amarelo: transmissão envolve devedor ou pessoa do grupo econômico
    - vermelho: alienação PELO DEVEDOR depois do ajuizamento

    O vermelho sinaliza fraude à execução, e fraude à execução pressupõe alienação
    feita PELO EXECUTADO (CPC/2015, art. 792). Antes bastava a data ser posterior ao
    ajuizamento, o que pintava de vermelho transmissão de pessoa do grupo e até de
    terceiro sem nenhuma relação com a execução — nenhuma das duas caracteriza a fraude.
    Transmissão de pessoa do grupo continua merecendo atenção, e por isso segue no
    amarelo; mas amarelo e vermelho não são a mesma tese.
    """
    alertas = set()
    todas_partes = list(devedores or []) + list(relacionados or [])
    transmissoes = resultado.get("transmissoes_estruturadas") or []

    for t in transmissoes:
        if isinstance(t, dict):
            de_doc    = t.get("de_doc") or ""
            para_doc  = t.get("para_doc") or ""
            de_nome   = t.get("de_nome") or ""
            para_nome = t.get("para_nome") or ""
            data_str  = t.get("data") or ""
        else:
            de_doc    = getattr(t, "de_doc", "") or ""
            para_doc  = getattr(t, "para_doc", "") or ""
            de_nome   = getattr(t, "de_nome", "") or ""
            para_nome = getattr(t, "para_nome", "") or ""
            data_str  = getattr(t, "data", "") or ""

        if todas_partes and (
            _mat_parte_envolvida(todas_partes, de_doc, de_nome)
            or _mat_parte_envolvida(todas_partes, para_doc, para_nome)
        ):
            alertas.add("amarelo")

        # Só quem TRANSMITIU importa aqui, e só se for devedor: quem adquire não
        # frauda a própria execução, e o grupo não é parte dela.
        alienou_devedor = bool(devedores) and _mat_parte_envolvida(devedores, de_doc, de_nome)
        if alienou_devedor and data_ajuizamento and data_str:
            try:
                p = data_str.strip().split("/")
                dt = date(int(p[2]), int(p[1]), int(p[0]))
                if dt >= data_ajuizamento:
                    alertas.add("vermelho")
            except Exception:
                pass

    return alertas


# ── Análise de PDF ────────────────────────────────────────────────────────────

def _mat_erro_troca_chave(exc: Exception) -> bool:
    return _erro_gemini_permite_failover(exc)


def _mat_consolidar_extracao(extracao: str, clients: list) -> dict:
    cfg = types.GenerateContentConfig(
        response_mime_type="application/json",
        response_schema=MatriculaExtraida,
        temperature=0,
        max_output_tokens=65536,
    )
    prompt = (
        PROMPT_MATRICULA
        + "\n\nEXTRACAO FACTUAL COMPLETA DA MATRICULA:\n"
        + extracao
        + "\n\nAntes de responder, faça três conferencias finais: "
        "(i) proprietário e frações refletem a última transmissão; "
        "(ii) todo ônus está em vigentes ou cancelados; "
        "(iii) matrícula encerrada/baixada e matrículas derivadas aparecem no primeiro campo."
    )

    ultimo_erro = None
    for pos, client in enumerate(clients):
        try:
            def _call():
                response = client.models.generate_content(
                    model=GEMINI_MODEL_RELATORIO,
                    contents=[prompt],
                    config=cfg,
                )
                return MatriculaExtraida.model_validate_json(response.text).model_dump()

            return _retry(_call, tentativas=3, espera_base=10)
        except Exception as exc:
            ultimo_erro = exc
            if not _mat_erro_troca_chave(exc) or pos == len(clients) - 1:
                raise
    raise ultimo_erro or RuntimeError("Nenhum cliente Gemini disponível para consolidação.")


def _mat_analisar_pdf(caminho_pdf: str, client_extracao, clients_consolidacao: list) -> dict:
    paginas_digitalizadas = _paginas_digitalizadas_pdf(caminho_pdf)
    model_extracao = GEMINI_MODEL_OCR if paginas_digitalizadas else GEMINI_MODEL_EXTRACAO
    contexto_paginas = (
        "\n\nCONTEXTO TÉCNICO DA FONTE — NÃO COPIAR COMO REFERÊNCIA:\n"
        f"- arquivo original: {Path(caminho_pdf).name}\n"
        "- este arquivo foi enviado integralmente; a página local 1 é a página absoluta 1 do PDF\n"
        "- páginas digitalizadas detectadas: "
        + (", ".join(str(p) for p in paginas_digitalizadas) if paginas_digitalizadas else "nenhuma")
        + "\nUse a numeração apenas para vincular cada R./AV. à página absoluta correta."
    )
    tmp_ascii_path = None
    try:
        caminho_pdf.encode("ascii")
        upload_path = caminho_pdf
    except UnicodeEncodeError:
        import shutil as _shutil
        tmp_ascii = tempfile.NamedTemporaryFile(suffix=".pdf", delete=False)
        tmp_ascii.close()
        _shutil.copy2(caminho_pdf, tmp_ascii.name)
        upload_path = tmp_ascii.name
        tmp_ascii_path = tmp_ascii.name

    arquivo = None
    try:
        arquivo = client_extracao.files.upload(file=upload_path)

        # Aguarda processamento do arquivo
        total_wait, wait_time = 0, 1
        while total_wait < 120:
            state_name = getattr(getattr(arquivo, "state", None), "name", "")
            if state_name in ("ACTIVE", "FAILED"):
                break
            time.sleep(wait_time)
            total_wait += wait_time
            arquivo = client_extracao.files.get(name=arquivo.name)
            wait_time = min(wait_time + 1, 4)
        if getattr(getattr(arquivo, "state", None), "name", "") == "FAILED":
            raise RuntimeError("O Gemini não conseguiu processar o PDF da matrícula.")

        cfg_extracao = types.GenerateContentConfig(
            temperature=0,
            max_output_tokens=65536,
        )

        def _extrair():
            response = client_extracao.models.generate_content(
                model=model_extracao,
                contents=[PROMPT_EXTRACAO_MATRICULA + contexto_paginas, arquivo],
                config=cfg_extracao,
            )
            if not response.text or not response.text.strip():
                raise RuntimeError("O modelo de extração devolveu conteúdo vazio.")
            return response.text

        extracao = _retry(_extrair, tentativas=3, espera_base=10)
        resultado = _mat_consolidar_extracao(extracao, clients_consolidacao)
        return _mat_pos_processar(resultado)
    finally:
        if arquivo is not None:
            try:
                client_extracao.files.delete(name=arquivo.name)
            except Exception:
                pass
        if tmp_ascii_path:
            try:
                os.remove(tmp_ascii_path)
            except Exception:
                pass


def _mat_worker(args) -> tuple:
    idx, pdf_path, clients = args
    ultimo_erro = None
    for tentativa in range(len(clients)):
        client_idx = (idx + tentativa) % len(clients)
        client_extracao = clients[client_idx]
        clients_consolidacao = clients[client_idx:] + clients[:client_idx]
        try:
            linha = _mat_analisar_pdf(pdf_path, client_extracao, clients_consolidacao)
            return idx, linha, None
        except Exception as exc:
            ultimo_erro = exc
            if not _mat_erro_troca_chave(exc) or tentativa == len(clients) - 1:
                break
    return idx, None, str(ultimo_erro)


# ── Helpers de nome ───────────────────────────────────────────────────────────

def _mat_limpar_nome(nome):
    nome = unicodedata.normalize("NFKD", nome)
    nome = nome.encode("ascii", "ignore").decode("ascii")
    nome = re.sub(r"[^A-Za-z0-9._-]", "_", nome)
    return nome


# ── Geração do Excel ──────────────────────────────────────────────────────────

COLUNAS_RENAME_MAT = {
    "matricula": "Matrícula", "comarca": "Comarca",
    "proprietario_atual": "Proprietário Atual", "descricao_imovel": "Descrição do Imóvel",
    "fracao_ideal": "Fração Ideal",
    "transmissoes_averbadas_registradas": "Transmissões",
    "onus_vigentes_registrados_averbados": "Ônus Vigentes",
    "principais_garantias_vigentes": "Principais Garantias Vigentes",
    "observacoes": "Observações", "onus_cancelados": "Ônus Cancelados",
    "grau_confianca": "Grau de Confiança",
}
COLUNAS_ORDEM_MAT = [
    "Matrícula", "Comarca", "Proprietário Atual", "Descrição do Imóvel",
    "Transmissões", "Ônus Vigentes", "Principais Garantias Vigentes",
    "Observações", "Ônus Cancelados",
    "Fração Ideal", "Valor da Avaliação Definitiva (VM)",
    "Valor da Avaliação Definitiva (VP)", "Valor Total do Ônus",
    "Saldo Avaliação - Ônus", "Grau de Confiança",
]
LARGURAS_MAT = {1:12, 2:22, 3:30, 4:45, 5:55, 6:55, 7:35, 8:50, 9:50, 10:14, 11:22, 12:22, 13:22, 14:22, 15:18}
COL_VM, COL_ONUS, COL_SALDO = 11, 13, 14

FILL_AMARELO  = PatternFill("solid", fgColor="FFFACD")  # lemon chiffon
FILL_VERMELHO = PatternFill("solid", fgColor="FFD0D0")  # rosa claro
FILL_PAR      = PatternFill("solid", fgColor="F2F7FB")
FILL_IMPAR    = PatternFill("solid", fgColor="FFFFFF")


def _mat_borda(estilo="thin", cor="BFBFBF"):
    s = Side(style=estilo, color=cor)
    return Border(left=s, right=s, top=s, bottom=s)


def _mat_estimar_linhas(texto, larg):
    if not texto: return 1
    chars = max(int(larg * 1.3), 1)
    total = 0
    for bloco in str(texto).split("\n"):
        total += 1 if bloco.strip() == "" else math.ceil(max(len(bloco), 1) / chars)
    return max(total, 1)


def _mat_salvar_excel(df, caminho, alert_map: dict):
    """
    alert_map: {df_index (0-based): set de alertas ('amarelo', 'vermelho')}
    """
    df.to_excel(caminho, index=False)
    wb = load_workbook(caminho)
    ws = wb.active
    ws.title = "Matriculas"
    n = len(df)
    ultima = 1 + n
    total  = ultima + 1
    lvm    = get_column_letter(COL_VM)
    lonus  = get_column_letter(COL_ONUS)

    for row in range(2, ultima + 1):
        ws.cell(row, COL_SALDO).value = f"={lvm}{row}-{lonus}{row}"
    ws.cell(total, 1).value = "TOTAL"
    for col in [COL_VM, COL_ONUS, COL_SALDO]:
        letra = get_column_letter(col)
        ws.cell(total, col).value = f"=SUM({letra}2:{letra}{ultima})"

    for col_idx, larg in LARGURAS_MAT.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = larg

    header_fill = PatternFill("solid", fgColor="1F4E79")
    ws.row_dimensions[1].height = 36
    for cell in ws[1]:
        cell.fill      = header_fill
        cell.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        cell.border    = _mat_borda("medium", "FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    colunas_center = {1, 2, 10, 11, 12, 13, 14, 15}
    for row_idx in range(2, ultima + 1):
        df_idx = row_idx - 2        # índice 0-based no DataFrame
        alertas = alert_map.get(df_idx, set())

        if "vermelho" in alertas:
            fill = FILL_VERMELHO
        elif "amarelo" in alertas:
            fill = FILL_AMARELO
        else:
            fill = FILL_PAR if row_idx % 2 == 0 else FILL_IMPAR

        max_linhas = 1
        for col_idx in range(1, ws.max_column + 1):
            cell  = ws.cell(row_idx, col_idx)
            horiz = "center" if col_idx in colunas_center else "left"
            cell.font      = Font(name="Calibri", size=10, color="1A1A1A")
            cell.fill      = fill
            cell.border    = _mat_borda("thin", "D0D0D0")
            cell.alignment = Alignment(horizontal=horiz, vertical="top", wrap_text=True)
            nl = _mat_estimar_linhas(cell.value, LARGURAS_MAT.get(col_idx, 20))
            max_linhas = max(max_linhas, nl)
        ws.row_dimensions[row_idx].height = min(max(16 * max_linhas + 6, 24), 409)

    total_fill = PatternFill("solid", fgColor="D6E4F0")
    ws.row_dimensions[total].height = 24
    for cell in ws[total]:
        cell.fill      = total_fill
        cell.font      = Font(bold=True, name="Calibri", size=10, color="1F4E79")
        cell.border    = _mat_borda("medium", "9DC3E6")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

    fmt_brl = r'R$\ #,##0.00'
    for col in [COL_VM, COL_VM + 1, COL_ONUS, COL_SALDO]:
        for row in range(2, total + 1):
            ws.cell(row, col).number_format = fmt_brl

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ultima}"

    # Legenda de cores (linha após o total)
    leg_row = total + 2
    ws.cell(leg_row, 1).value = "Legenda:"
    ws.cell(leg_row, 1).font = Font(bold=True, name="Calibri", size=9)
    ws.cell(leg_row, 2).fill = FILL_AMARELO
    ws.cell(leg_row, 2).value = "Transmissão envolvendo devedor ou grupo econômico"
    ws.cell(leg_row, 2).font = Font(name="Calibri", size=9)
    ws.cell(leg_row, 3).fill = FILL_VERMELHO
    ws.cell(leg_row, 3).value = "Transmissão após ajuizamento da execução"
    ws.cell(leg_row, 3).font = Font(name="Calibri", size=9)

    wb.save(caminho)


# ── Função pública ────────────────────────────────────────────────────────────

def _mat_pares_dos_campos(campos: tuple) -> tuple:
    """Reagrupa os campos achatados da interface em (devedores, pessoas do grupo).

    A UI manda quatro blocos do mesmo tamanho, nesta ordem: nomes de devedor, docs de
    devedor, nomes do grupo, docs do grupo. Mesmo padrão da aba de RJ, que achata
    nomes+docs e reparea do lado Python.
    """
    if not campos:
        return [], []
    bloco = len(campos) // 4
    if bloco == 0:
        return [], []
    nomes_dev, docs_dev = campos[:bloco], campos[bloco:2 * bloco]
    nomes_grp, docs_grp = campos[2 * bloco:3 * bloco], campos[3 * bloco:4 * bloco]
    return (
        list(zip(nomes_dev, docs_dev)),
        list(zip(nomes_grp, docs_grp)),
    )


def mat_gerar_excel(arquivos, data_ajuizamento: str = "", *campos):
    if not arquivos:
        yield "Envie ao menos um PDF.", "", None
        return

    # Parse dos parâmetros opcionais
    pares_devedores, pares_grupo = _mat_pares_dos_campos(campos)
    devedores  = _mat_parse_parties(pares_devedores)
    relacionados = _mat_parse_parties(pares_grupo)
    data_ajuiz: date | None = None
    if data_ajuizamento and data_ajuizamento.strip():
        try:
            p = data_ajuizamento.strip().split("/")
            data_ajuiz = date(int(p[2]), int(p[1]), int(p[0]))
        except Exception:
            pass

    usar_alertas = bool(devedores or relacionados or data_ajuiz)

    try:
        clients = _mat_get_clients()
    except Exception as e:
        yield f"Erro de configuração: {e}", "", None
        return

    # Limpa arquivos temporários da execução anterior
    for f in os.listdir("tmp_pdfs"):
        try: os.remove(os.path.join("tmp_pdfs", f))
        except Exception: pass

    pdfs = []
    for arq in arquivos:
        origem = getattr(arq, "name", arq)
        nome   = _mat_limpar_nome(os.path.basename(origem))
        destino = os.path.join("tmp_pdfs", nome)
        with open(origem, "rb") as o, open(destino, "wb") as d:
            d.write(o.read())
        pdfs.append(destino)

    n = len(pdfs)
    workers = min(n, len(clients), MATRICULAS_MAX_WORKERS)
    log = [
        f"Analisando {n} matricula(s) em paralelo ({workers} worker(s))...",
        f"   Leitura pesquisável: {GEMINI_MODEL_EXTRACAO} | Leitura digitalizada: {GEMINI_MODEL_OCR} | "
        f"Consolidacao: {GEMINI_MODEL_RELATORIO} | "
        f"{len(clients)} chave(s) com failover",
    ]
    if usar_alertas:
        partes = []
        if data_ajuiz:
            partes.append(f"ajuizamento: {data_ajuizamento.strip()}")
        if devedores:
            partes.append(f"{len(devedores)} devedor(es)")
        if relacionados:
            partes.append(f"{len(relacionados)} pessoa(s) do grupo")
        log.append(f"   Alertas ativos: {', '.join(partes)}")
    yield "\n".join(log), "", None

    resultados_dict: dict = {}
    erros_dict: dict     = {}

    # Cada worker começa por uma chave diferente e pode alternar para as demais
    # se houver falha de autenticação, cota ou disponibilidade do modelo.
    args_list = [(i, pdf, clients) for i, pdf in enumerate(pdfs)]

    with concurrent.futures.ThreadPoolExecutor(max_workers=workers) as ex:
        futures = {ex.submit(_mat_worker, args): args[0] for args in args_list}
        concluidos = 0
        for future in concurrent.futures.as_completed(futures):
            idx, linha, erro = future.result()
            concluidos += 1
            nome_pdf = Path(pdfs[idx]).name
            if erro:
                erros_dict[idx] = erro
                log.append(f"   [{concluidos}/{n}] {nome_pdf} — Erro: {erro}")
            else:
                resultados_dict[idx] = linha
                log.append(f"   [{concluidos}/{n}] {nome_pdf} — OK")
            yield "\n".join(log), "", None

    # Ordena por índice original
    indices_ok = sorted(resultados_dict.keys())
    hoje = date.today()
    alert_map: dict = {}

    resultados_finais = []
    for pos, idx in enumerate(indices_ok):
        r = resultados_dict[idx]
        onus_list = r.pop("onus_financeiros", None) or []
        onus_vigentes_txt = r.get("onus_vigentes_registrados_averbados", "") or ""
        r["valor_total_onus_calculado"] = _mat_calcular_valor_onus(
            onus_list, onus_vigentes_texto=onus_vigentes_txt, data_referencia=hoje
        )
        r["principais_garantias_vigentes"] = _mat_resumo_garantias(onus_list)

        if usar_alertas:
            alertas = _mat_detectar_alertas(r, devedores, relacionados, data_ajuiz)
            if alertas:
                alert_map[pos] = alertas

        r.pop("transmissoes_estruturadas", None)  # não vai para o Excel
        resultados_finais.append(r)

    if not resultados_finais:
        log.append("\nNenhuma matrícula processada com sucesso.")
        yield "\n".join(log), "Nenhum resultado.", None
        return

    df = pd.DataFrame(resultados_finais)
    df = df.rename(columns=COLUNAS_RENAME_MAT)
    if "Fração Ideal" not in df.columns:
        df["Fração Ideal"] = ""
    df["Valor da Avaliação Definitiva (VM)"] = ""
    df["Valor da Avaliação Definitiva (VP)"] = ""
    df["Valor Total do Ônus"] = df.pop("valor_total_onus_calculado") if "valor_total_onus_calculado" in df.columns else ""
    df["Saldo Avaliação - Ônus"] = ""
    df = df[[c for c in COLUNAS_ORDEM_MAT if c in df.columns]]

    caminho = os.path.join("resultados", "resultado_matriculas.xlsx")
    _mat_salvar_excel(df, caminho, alert_map)

    linhas_alerta = []
    if alert_map:
        n_am = sum(1 for a in alert_map.values() if "amarelo" in a and "vermelho" not in a)
        n_vm = sum(1 for a in alert_map.values() if "vermelho" in a)
        if n_am: linhas_alerta.append(f"{n_am} matrícula(s) amarela(s)")
        if n_vm: linhas_alerta.append(f"{n_vm} matrícula(s) vermelha(s)")

    resumo = f"{len(resultados_finais)}/{n} OK"
    if linhas_alerta:
        resumo += " | " + ", ".join(linhas_alerta)
    if erros_dict:
        resumo += f" | {len(erros_dict)} erro(s)"

    log.append(f"\nExcel gerado — {resumo}")
    yield "\n".join(log), "Excel pronto!", caminho


def mat_responder(pergunta: str, log_texto: str):
    try:
        clients = _mat_get_clients()
        ultimo_erro = None
        for pos, client in enumerate(clients):
            try:
                return _responder_pergunta_generica(
                    pergunta, log_texto, client, GEMINI_MODEL_QA
                )
            except Exception as exc:
                ultimo_erro = exc
                if not _mat_erro_troca_chave(exc) or pos == len(clients) - 1:
                    raise
        raise ultimo_erro or RuntimeError("Nenhum cliente Gemini disponível.")
    except Exception as e:
        return f"Erro: {e}"
